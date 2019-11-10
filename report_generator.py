import screener_handler
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import pandas as pd
import logging
import csv
import os
import openpyxl
import xlwings as xw
import excel2img

#LOGGING SETTINGS:
logging.basicConfig(level=logging.DEBUG)

#GLOBAL VARIABLES
TEMPLATE_PATH = 'data/Template.xlsx'
REPORT_FORMAT = '.png'
TEMPLATE_URL1 = 'https://www.nasdaqbaltic.com/statistics/en/shares?date=yyyy.mm.dd'
DATE_OF_TODAY = datetime.now()
DATE_OF_LAST_WEEK = datetime.now() - timedelta(7)

#PUBLIC FUNCTIONS
def extract_company_name(td_tag_within_company_row):
    company_name = td_tag_within_company_row[0].a.text.strip()
    return company_name

def extract_company_ticker(td_tag_within_company_row, company_name):
    temp_ticker = td_tag_within_company_row[0].text
    ticker = temp_ticker.replace('\t' + company_name, '').replace('\tLP', '').replace('\t', '').replace('-', '').replace('!', '').strip()
    return ticker

def extract_last_price(td_tag_within_company_row):
    last_price = td_tag_within_company_row[6].text
    return last_price

def subtract_day(date):
    shifted_date = date - timedelta(1)
    return shifted_date

class InvestmentReport:
    def __init__(self):
        pass

    def server_response_checker(self):
        '''method returns server response code of nasdaqomxbaltic.com website.'''
        nasdaqomxbaltic_home_url = 'https://www.nasdaqbaltic.com/?lang=en'
        r_check = requests.get(nasdaqomxbaltic_home_url)
        self.server_response = r_check.status_code

    def trading_days_checker(self):
        '''method checks if nasdaq website provides trading data for desired dates
        if not - iterates until combination of dates is found, that provides desired data for report'''      
        self.date_of_today = DATE_OF_TODAY
        self.date_of_last_week = DATE_OF_LAST_WEEK
        self.url_builder()
        iteration = 0
        while self.soups_contain_errors(self.url_prices_today, self.url_prices_last_week) == False: 
            iteration += 1
            self.date_of_today = subtract_day(self.date_of_today)
            self.date_of_last_week = subtract_day(self.date_of_last_week)
            self.url_builder()
            logging.debug('URL checking iteration no: ' + str(iteration))
        logging.debug('Found working URLs at interation no: ' + str(iteration))
        logging.debug('Program proceeds with these URLs: ' + self.url_prices_today + ' and ' + self.url_prices_last_week)
 
    def url_builder(self):
        '''Prepares two url's of current date and last week's 
        nasdaqomxbaltic trading session equity list prices. Url's will be used for testing & scraping'''
        #string preparation
        self.date_of_today_string = self.date_of_today.strftime('%Y.%m.%d')
        self.last_week_date_string = self.date_of_last_week.strftime('%Y.%m.%d')
        # two output URL's:
        self.url_prices_today = TEMPLATE_URL1.replace('yyyy.mm.dd', self.date_of_today_string)
        self.url_prices_last_week = TEMPLATE_URL1.replace('yyyy.mm.dd', self.last_week_date_string)

    def soups_contain_errors(self, url1, url2):
        '''Takes two passed urls (as string arguments), checks and returns True if both soups contain
        desired data for report generation. Returned Boolean is evaluated in trading_days_checker() method'''
        r1 = requests.get(url1)
        r2 = requests.get(url2)
        soup1 = BeautifulSoup(r1.text, features='lxml')
        soup2 = BeautifulSoup(r2.text, features='lxml')
        # Non-trading days/weekends soup contain unique div with class 'col-12 text-info'
        if soup1.find('div', class_='col-12 text-info') == None and soup2.find('div', class_='col-12 text-info') == None:
            logging.debug('Both provided urls contain data, allowing to proceed with the program')
            return True
        else:
            logging.debug('One these URLs was invalid for data collection: ' + url1 + ' or ' + url1)
            return False
    
    def get_prices_soup(self, url):
        '''Creates a soup from url'''
        r = requests.get(url)
        self.soup = BeautifulSoup(r.text, features='lxml')

    def get_rows_containing_data(self):
        '''Extracts rows containing actual companies data (tr's have 27 td's and contains child class 'text16 compname')'''
        rows_containers = self.soup.find_all('tr')
        self.companies_rows = []
        for tr in rows_containers:
            if len(tr) == 27 and 'text16 compname' in str(tr):
                self.companies_rows.append(tr)
        
    def get_scrape_results(self):
        '''Iterates through row elements and outputs a list of dictionaries for each company.'''
        #General list of dictionaries for all companies:
        self.scrape_output = []
        #iterating through companies to collect data:
        for company in self.companies_rows:
            temp_d = {}
            td_tag_within_company_row = company.findAll('td')
            temp_d['Name'] = extract_company_name(td_tag_within_company_row)
            temp_d['Ticker'] = extract_company_ticker(td_tag_within_company_row, temp_d['Name'])
            temp_d['Last Price'] = extract_last_price(td_tag_within_company_row)
            # output list of dictionaries:
            self.scrape_output.append(temp_d)

    def csv_filename(self, url):
        '''Prepares appropriate csv file names, depending on what url was passed'''
        self.today_csv_filename = 'data/Prices_' + self.date_of_today_string + '.csv'
        self.last_week_csv_filename = 'data/Prices_' + self.last_week_date_string + '.csv'
        #csv filename depends on URL being used for scraping:
        if self.date_of_today_string in url:
            self.temp_csv_filename = self.today_csv_filename
        elif self.last_week_date_string in url:
            self.temp_csv_filename = self.last_week_csv_filename
        else:
            logging.debug('Faulty URL passed to function')

    def export_scrape_results_csv(self):
        keys = self.scrape_output[0].keys()
        with open (self.temp_csv_filename, 'w', newline='') as outputfile:
            dict_writer = csv.DictWriter(outputfile, keys)
            dict_writer.writeheader()
            dict_writer.writerows(self.scrape_output)

    def scrape_to_csv(self, url):
        '''List of methods to be performed to scrape and output desired data to csv'''
        logging.debug('Executing scrape to csv(url)')
        self.get_prices_soup(url)
        self.get_rows_containing_data()
        self.get_scrape_results()
        self.csv_filename(url)
        self.export_scrape_results_csv()
        logging.debug('Scrape and output of method scrape_to_csv(url) successfully executed')

    def form_joint_dataframe(self):
        '''Takes two csv files, cleans data and merges to joint dataframe. Calculates additional column ['Change, %']'''
        last_week_df = pd.read_csv(self.last_week_csv_filename, header = 0, encoding = 'ISO-8859-1', names=['Name', 'Ticker', 'Last Week Price'])
        today_df = pd.read_csv(self.today_csv_filename, header = 0, encoding = 'ISO-8859-1', names=['Name', 'Ticker', 'Last Price'])
        #trimming dataframes:
        data_from_last_week_df = last_week_df[['Ticker', 'Last Week Price']]
        data_from_today_df = today_df[['Ticker', 'Last Price']]
        # joining dataframes and making a relative price change calulation:
        self.joint_df = pd.merge(data_from_last_week_df, data_from_today_df, on='Ticker')
        self.joint_df['Change, %'] = (self.joint_df['Last Price'] - self.joint_df['Last Week Price'])/self.joint_df['Last Week Price']
        
    def get_best_worst_performers_df(self):
        '''Forms two dataframes of best and worst companies by price performance, sorted accordingly'''
        self.top_performers = self.joint_df.sort_values(by='Change, %', ascending=False).drop(columns = ['Last Week Price', 'Last Price']).head(5)
        self.worst_performers = self.joint_df.sort_values(by='Change, %').drop(columns = ['Last Week Price', 'Last Price']).head(5)

    def get_data_from_screener(self, csvfilename):
        '''Takes argument of csvfile path and writes dataframe returned by screener_handler.py module\'s method run as variable'''
        imported_list = screener_handler.Screener_instance.run(self.today_csv_filename)
        self.random_ratio = imported_list[0]
        self.random_bool = imported_list[1]
        self.df_from_screener = imported_list[2]
        
    def load_data_to_template_excel(self):
        '''Loads dataframes, variables to pre-made Template.xlsx, without modifying the rest
        of the document. Saves changes'''
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb['Main']
        ws['B18'] = self.date_of_today_string
        ws['B19'] = self.last_week_date_string
        ws['A21'] = self.random_ratio
        ws['B22'] = self.random_bool
        with pd.ExcelWriter(TEMPLATE_PATH, engine='openpyxl') as writer:
            writer.book = wb
            writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
            self.top_performers.to_excel(writer, index=False, header = False, startrow = 1, sheet_name = 'Main')
            self.worst_performers.to_excel(writer, index=False, header = False, startrow = 7, sheet_name = 'Main')
            self.df_from_screener.to_excel(writer, index=False, header = False, startrow = 23, startcol = 20,  sheet_name = 'Main')
            writer.save()

    def remove_chart_outline(self):
        '''Copies template sheet to xlsm, removes charts outline and changes axis number formatting via VBA script'''
        self.temp_TEMPLATE_PATH = TEMPLATE_PATH.replace('xlsx', 'xlsm')
        #Creating temp xlsxm empty workbook + macro
        writer = pd.ExcelWriter(TEMPLATE_PATH, engine='xlsxwriter')
        wb = writer.book
        wb.filename = self.temp_TEMPLATE_PATH
        wb.add_vba_project('data/vbaProject.bin')
        writer.save()
        #Opens temp xlsm workbook and executes VBA script to copy data, formats, charts and REMOVE CHART OUTLINE 
        wb = xw.Book(self.temp_TEMPLATE_PATH)
        app = wb.app
        macro_vba = app.macro('Loader')
        macro_vba()
        wb.save()
        wb.close()
        app.quit()

    def generate_output(self):
        '''Generates desired image file from Template's named range'''
        self.report_file_name = 'Report '+self.date_of_today_string + REPORT_FORMAT
        #Generate output image - report:
        excel2img.export_img(self.temp_TEMPLATE_PATH, 'data/' + self.report_file_name, None, 'Output_Area')

    def clean_temp_files(self):
        '''Cleans unneccessary files after program has finished'''
        os.remove(screener_handler.SCREENER_FILENAME_PATH)
        os.remove(self.today_csv_filename)
        os.remove(self.last_week_csv_filename)
        os.remove(self.temp_TEMPLATE_PATH)


    def run(self):
        self.server_response_checker()
        if self.server_response == 200:
            self.trading_days_checker()
            logging.debug('Found a set of trading days; built URLs for scrapping')
            self.scrape_to_csv(self.url_prices_today)
            logging.debug(f'About to scrape another URL: {self.url_prices_last_week}')
            self.scrape_to_csv(self.url_prices_last_week)
            logging.debug('2 csv files were created in /data folder')
            self.form_joint_dataframe()
            logging.debug('Joint dataframe was created')
            self.get_best_worst_performers_df()
            logging.debug('Two dataframes of best and worst performing stocks formed')
            logging.debug(f'Querying Screener class, passing fresh stock prices: {self.today_csv_filename}')            
            self.get_data_from_screener(self.today_csv_filename)
            self.load_data_to_template_excel()
            logging.debug('All desired data loaded to Template.xlsx')
            self.remove_chart_outline()
            logging.debug('temp xlsm file ready to be passed for image processing')
            self.generate_output()
            logging.debug(f'Report named: {self.report_file_name} is created in data/ folder')
            self.clean_temp_files()
            logging.debug('Temporary files have been deleted.\n---------------Program finished executing---------------')
        else:
            logging.warning('\nWebsite is currently unreachable;\nProgram has terminated.')


if __name__ == '__main__':
    InvestmentReport().run()