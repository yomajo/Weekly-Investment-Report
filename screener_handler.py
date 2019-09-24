import logging
import os
import sys
from shutil import copy
import openpyxl
import csv
import random 
import xlwings as xw
import pandas as pd

#LOGGING SETTINGS:
logging.basicConfig(level=logging.DEBUG)

#GLOBAL VARIABLES
#Note, that due to sensitive information inside screener it is not publicly available
SCREENER_ORIGINAL_ABS_PATH = 'C:/Svarbu/OMXB analitika.xlsm'
SCREENER_FILENAME = SCREENER_ORIGINAL_ABS_PATH.split('/')[-1]
SCREENER_FILENAME_PATH = 'data/' + SCREENER_FILENAME
SRC_DIR = '/'.join(SCREENER_ORIGINAL_ABS_PATH.split('/')[:-1])

class Screener:
    def __init__(self):
        pass
    
    def screener_exists(self, abs_path):
        '''Checks if screener file exists and returns True if it does. Takes one arg - absolute path to file of interest'''
        if os.path.exists(abs_path):
            return True

    def make_temp_screener_copy(self, src_path):
        '''Creates a temporary copy of file to project /data folder. Takes an input of source path as string'''
        copy(src_path, os.getcwd() + '/data/')

    def last_price_csv_to_dict(self, csvfilename):
        '''Takes argument of csv filename as a string. Method creates dictionary of company ticker as key and last price as value'''
        with open(csvfilename) as csv_file:
            csv_reader = csv.reader(csv_file)
            #skipping header row
            next (csv_reader)
            self.ticker_price_dict={}
            for csv_line in csv_reader:
                try:
                    self.ticker_price_dict[csv_line[1]] = float(csv_line[2])
                except:
                    self.ticker_price_dict[csv_line[1]] = csv_line[2]
    
    def open_screener(self, rel_path):
        '''Trying to manipulate fresh copy of xlsm file'''
        logging.debug('Opening worksheet {}'.format(rel_path))
        # Limitation of openpyxl about loosing formulas or being unable to read cell contents where formulas are present
        # Workaround: open two sessions in different modes; one for reading, another for writing
        self.wb = openpyxl.load_workbook(rel_path, keep_vba=True)
        self.wb_read = openpyxl.load_workbook(rel_path, data_only=True, keep_vba=True)

    def load_new_prices(self):
        '''Editing inside copy of excel screener'''
        ws = self.wb['Prices']
        ws_read = self.wb_read['Prices']
        #Determining at which row should price uploading begin:
        for cell in ws['L']:
            if cell.value == 'Last Price':
                self.input_range_row = cell.row
                break

        self.count_uploaded = 0
        #iterating excel ticker range, if a match in csv formed dict is found - value is updated
        for cell in range(self.input_range_row, ws.max_row):
            if ws_read[f'W{cell}'].value in self.ticker_price_dict:
                ws[f'L{cell}'].value = self.ticker_price_dict[ws_read[f'W{cell}'].value]
            self.count_uploaded += 1
        logging.debug(f'A total number of {self.count_uploaded} companies last prices were uploaded to temp. excel file')

    def use_latest_values(self):
        '''Sets additional cell values, so most up to date data is used in screener file'''
        ws = self.wb['Summary']
        ws['C1'] = self.wb['Universals']['E18'].value
        ws['C2'] = ws['X1'].value
        logging.debug('Additinal cell values have been set in Summary sheet C1:C2 with values {} and {}'.format(self.wb['Universals']['E18'].value, ws['X1'].value))

    def get_available_ratios(self):
        '''Get a collection of ratios available in the screener as dict'''
        ws = self.wb_read['Summary']
        self.ratios_dict = {}
        # Starting location of ratios and their categories in Summary sheet defined by start_row_col tuple
        start_row_col = (4, 29)
        #Iterating through ratio categories (row=4), and ratios (columns 29-31) to collect a dict of lists for each category
        logging.debug('Looping through [Summary] sheet and collecting available ratios')        
        for col in range(start_row_col[1], ws.max_column):
                if ws.cell(row=start_row_col[0], column=col).value != None:
                        temp_key = ws.cell(row=start_row_col[0], column=col).value
                        ratios_within_category = []
                        # If outter column (ratio category) cell value is not emply - start another cycle of collecting ratios iterating through rows
                        for row in range(start_row_col[0] + 1, ws.max_row):
                                if ws.cell(row=row, column=col).value != None:
                                        ratios_within_category.append(ws.cell(row=row, column=col).value)
                                else:
                                        self.ratios_dict[temp_key] = ratios_within_category
                                        break             
                else:
                        break

    def pick_random_values_for_sorting(self):
        '''Outputs three randomly picked values: 1. ratio category 2. ratio 3. True/False boolean for accending/descending sorting'''
        self.random_ratio_category = random.choice(list(self.ratios_dict))
        self.random_ratio = random.choice(self.ratios_dict[self.random_ratio_category])
        self.random_boolean = random.choice([True, False])
        logging.debug(f'Random selection. Ratio category: {self.random_ratio_category}, ratio within category: {self.random_ratio}, ascending sorting: {self.random_boolean}')

    def get_current_screener_table_headers(self):
        '''Get list of current screener table headers'''
        ws = self.wb['Summary']
        self.current_headers_list = []
        for c in range(1, 18):
            self.current_headers_list.append(ws.cell(row = 4, column = c).value)

    def push_ratio_to_screener_table(self, ratio_category, ratio):
        '''Takes arguments of ratio category, ratio. If ratio is not in table headers already.
        Method writes it to screener table header, ensuring corresponding data is present before reading it'''
        ws = self.wb['Summary']
        self.get_current_screener_table_headers()
        # Iterating through ratio categories in table upper headings to write random ratio to headers IF it does already exist in headers list 
        if self.random_ratio not in self.current_headers_list:
            for col in range(5, ws.max_column):
                if ws.cell(row=3, column=col).value == ratio_category:
                    ws.cell(row=4, column=col).value = ratio
                    break
        else:
            logging.debug(f'Randomly picked ratio {self.random_ratio} already exists in the headers list, no additional values were pushed into workbook')
            pass

    def recalculate_workbook(self, file_path):
        '''Opening workbook with xlwings as hidden instance to evaluate the formulas'''
        app = xw.App(visible=False)
        book = xw.Book(file_path)
        book.save(file_path)
        book.close()
        app.quit()
    
    def table_to_df(self):
        '''Returns Pandas dataframe formed from screener table'''
        ws_read = self.wb_read['Summary']
        table_start_row_col = (4, 1)
        screener_table_data_rows = []
        for r in range(table_start_row_col[0], ws_read.max_row):
            screener_table_data_cols = []
            if ws_read[f'E{r}'].value != None:
                for c in range(table_start_row_col[1], 18):
                    screener_table_data_cols.append(ws_read.cell(row=r, column=c).value)
                screener_table_data_rows.append(screener_table_data_cols)
            else:
                break
        # Converting list to dataframe
        self.headers = screener_table_data_rows.pop(0)
        self.df = pd.DataFrame(screener_table_data_rows, columns=self.headers)
        
    def crop_sort_df(self):
        '''Crops, sorts and formats original dataframe to top/bottom by random ratio picked before'''    
        cropped_df = self.df[[self.headers[1], self.random_ratio]]
        self.sorted_cropped_df = cropped_df.sort_values(by=[self.random_ratio], ascending=self.random_boolean).head(5)
    
    def form_export_data_list(self):
        '''Forms a list of selected random ratio, selected boolean and a dataframe created by Screener class to be returned'''
        self.export_list = [self.random_ratio, self.random_boolean, self.sorted_cropped_df]
    
    def close_screener(self):
        '''Saves screener and closes workbook'''
        logging.debug('Saving & Closing...')
        self.wb.save(SCREENER_FILENAME_PATH)
        self.wb.close()
        self.wb_read.close()

    def run(self, csvfilename):
        if self.screener_exists(SCREENER_ORIGINAL_ABS_PATH) == True:
            logging.debug(f'Screener {SCREENER_FILENAME} found')
            self.make_temp_screener_copy(SCREENER_ORIGINAL_ABS_PATH)
            logging.debug(f'Screener copy was created: {SCREENER_FILENAME_PATH}')
            self.open_screener(SCREENER_FILENAME_PATH)
            self.last_price_csv_to_dict(csvfilename)
            logging.debug(f'Uploading prices from {csvfilename} to screener')
            self.load_new_prices()
            self.use_latest_values()
            self.get_available_ratios()
            self.pick_random_values_for_sorting()
            logging.debug('Getting current headers list')
            self.push_ratio_to_screener_table(self.random_ratio_category, self.random_ratio)
            logging.debug(f'Closing workbook {SCREENER_FILENAME}')
            self.close_screener()
            logging.debug(f'Opening, calculating and saving workbook {SCREENER_FILENAME} via hidden excel instance using xlwings')
            self.recalculate_workbook(SCREENER_FILENAME_PATH)
            logging.debug('Reloading workbook on openpyxl')
            self.open_screener(SCREENER_FILENAME_PATH)
            logging.debug(f'Collecting screener table to dataframe, sorting according to selected random ratio: {self.random_ratio}; cropping resulting df')
            self.table_to_df()
            self.crop_sort_df()
            logging.debug('Exporting data to a list')
            self.form_export_data_list()
            self.close_screener()
            logging.debug(f'{os.path.basename(__file__)} finished executing; returning list containing selected random ratio and dataframe')
            return self.export_list
        else:
            logging.warning(f' PROGRAM TERMINATED. \n Please check directory: {SRC_DIR} \n File named {SCREENER_FILENAME} was not found there.')
            sys.exit()
        
# Creating a class instance be to accessed from another module
Screener_instance = Screener()

if __name__ == '__main__':
    prices_csvfilename = input('Enter a full csv filename containing stock prices that will be uploaded to screener(example: data/Prices.csv): ')
    Screener().run(prices_csvfilename)